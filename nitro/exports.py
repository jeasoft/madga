"""
Django Nitro 0.8 - Export Mixin for CSV and Excel exports.

Usage::

    from nitro.exports import ExportMixin

    class PropertyListView(ExportMixin, CompanyMixin, NitroListView):
        export_fields = ['name', 'address', ('landlord.legal_name', 'Propietario'), 'status']
        export_filename = 'propiedades'
        export_brand_color = '09C6AF'

Template::

    {% nitro_export_buttons %}
"""

import csv
import io
from datetime import date, datetime

from django.http import HttpResponse
from django.utils import timezone


def resolve_field_value(obj, field_path):
    """Resolve a dotted field path: 'landlord.legal_name' -> obj.landlord.legal_name"""
    value = obj
    for attr in field_path.split('.'):
        if value is None:
            return ''
        if callable(value):
            value = value()
        value = getattr(value, attr, '')
    if callable(value):
        value = value()
    return value


def format_export_value(value):
    """Format a value for export (dates, None, etc.)."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return timezone.localtime(value).strftime('%Y-%m-%d %H:%M') if timezone.is_aware(value) else value.strftime('%Y-%m-%d %H:%M')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return str(value)


class ExportMixin:
    """
    Mixin for CSV and Excel exports on list views.

    Intercepts GET requests with ?export=csv or ?export=excel
    and returns the filtered queryset as a downloadable file.

    Attributes:
        export_fields: List of field paths or (field, label) tuples.
        export_filename: Base filename without extension.
        export_brand_color: Hex color for Excel header (without #).
    """

    export_fields = []
    export_filename = ''
    export_brand_color = '09C6AF'

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export', '').lower()
        if export_format == 'csv':
            return self.export_csv(request)
        elif export_format == 'excel':
            return self.export_excel(request)
        return super().get(request, *args, **kwargs)

    def get_export_fields(self):
        """Return list of (field_path, label) tuples."""
        fields = []
        for f in self.export_fields:
            if isinstance(f, (list, tuple)):
                # Convert lazy translation strings to str for openpyxl compatibility
                fields.append((f[0], str(f[1])))
            else:
                # Auto-generate label from field name
                label = f.replace('_', ' ').replace('.', ' - ').title()
                fields.append((f, label))
        return fields

    def get_export_queryset(self, request):
        """Get the queryset for export. Uses the same filters as the list view."""
        return self.get_filtered_queryset()

    def get_export_row(self, obj, fields):
        """Get a row of values for the given object."""
        row = []
        for field_path, _label in fields:
            value = resolve_field_value(obj, field_path)
            row.append(format_export_value(value))
        return row

    def get_export_filename(self):
        """Get the filename for the export."""
        if self.export_filename:
            return self.export_filename
        if self.model:
            return self.model._meta.verbose_name_plural.replace(' ', '_')
        return 'export'

    def export_csv(self, request):
        """Export filtered queryset as CSV."""
        fields = self.get_export_fields()
        filename = self.get_export_filename()
        queryset = self.get_export_queryset(request)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        response.write('\ufeff')  # BOM for Excel UTF-8 compatibility

        writer = csv.writer(response)
        writer.writerow([label for _field, label in fields])

        for obj in queryset:
            writer.writerow(self.get_export_row(obj, fields))

        return response

    def export_excel(self, request):
        """Export filtered queryset as Excel (.xlsx)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return HttpResponse('openpyxl is required for Excel export', status=500)

        fields = self.get_export_fields()
        filename = self.get_export_filename()
        queryset = self.get_export_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = filename[:31]  # Excel limit

        # Header styling
        header_fill = PatternFill(start_color=self.export_brand_color, end_color=self.export_brand_color, fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            bottom=Side(style='thin', color='CCCCCC'),
        )

        # Write headers
        for col_idx, (_field, label) in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, obj in enumerate(queryset, 2):
            row_data = self.get_export_row(obj, fields)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-width columns
        for col_idx, (_field, label) in enumerate(fields, 1):
            max_length = len(label)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 3, 50)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Write to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response
